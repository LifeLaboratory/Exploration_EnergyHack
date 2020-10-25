import openpyxl
from base.base_sql import Sql
from openpyxl.styles import Alignment, Font, PatternFill
from app.company_data.processor import *


def generate_pc_report(id_company):
    name_file = 'Отчет_ПС.xlsx'
    new_file = 'static/Отчет_ПС_new.xlsx'

    wb = openpyxl.load_workbook(name_file)
    sheet = wb.active
    cursor = 6
    # id_company = 1


    networks = get_networks(id_company)


    def add_netowrk_label(network, cursor):
        sheet.merge_cells(f'A{cursor}:O{cursor}')
        fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')
        sheet[f'A{cursor}'].alignment = Alignment(horizontal='center')
        sheet[f'A{cursor}'] = network.get("Название")
        sheet[f'A{cursor}'].fill = fill


    def count_trans_in_rec(rec):
        sql_query = """
    select
    count(1)
    from "Трансформатор"
    join "Подстанция" on "Подстанция".id = "Трансформатор"."Подстанция"
    join "РЭС" on "РЭС".id = "Подстанция"."РЭС"
    
        """
        try:
            result = Sql.exec(query=sql_query)[0]['count']
        except:
            result = 0
        return result


    def add_rec(network, cursor):
        recs = get_rec(network.get('id'))
        for rec in recs:
            sheet.merge_cells(f'A{cursor}:A{cursor + count_trans_in_rec(rec)}')
            sheet[f'A{cursor}'].alignment = Alignment(horizontal='center', wrapText=True)
            sheet[f'A{cursor}'] = rec.get("Название")


    def get_pc(id_network):
        sql_query = f"""
    select
    *
    from "Подстанция"
    where "Сеть" = {id_network}
    order by id
        """
        return Sql.exec(query=sql_query)


    def add_pc(network, cursor):
        pcs = get_pc(network.get('id'))
        new_cursor = cursor
        for pc in pcs:
            cursor = new_cursor
            transs = get_trans(pc.get("id"))
            for trans in transs:
                sheet[f'F{new_cursor}'] = trans.get("id")
                sheet[f'G{new_cursor}'] = trans.get("Тип")
                sheet[f'H{new_cursor}'] = trans.get("Номинальная_Мощность")
                sheet[f'J{new_cursor}'] = trans.get("Год_изготовления")
                sheet[f'K{new_cursor}'] = trans.get("Год_включения")
                sheet[f'M{new_cursor}'] = trans.get("Техническое_состояние")
                new_cursor += 1
            sheet[f'B{cursor}'] = pc.get("id")
            sheet[f'C{cursor}'] = pc.get("Название")
            sheet[f'D{cursor}'] = pc.get("Класс_напряжения")
            sheet[f'E{cursor}'] = pc.get("Год_ввода")
            if new_cursor-1 < cursor:
                new_cursor += 1
            sheet.merge_cells(f'B{cursor}:B{new_cursor-1}')
            sheet.merge_cells(f'C{cursor}:C{new_cursor-1}')
            sheet.merge_cells(f'D{cursor}:D{new_cursor-1}')
            sheet.merge_cells(f'E{cursor}:E{new_cursor-1}')
            cursor += 1
        return new_cursor


    for network in networks:
        add_netowrk_label(network, cursor)
        cursor += 1
        cursor = add_pc(network, cursor)
        # cursor += 1
    # add_rec(network)


    wb.save(new_file)

