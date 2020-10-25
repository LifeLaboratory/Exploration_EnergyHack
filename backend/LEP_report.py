import openpyxl
from base.base_sql import Sql
from openpyxl.styles import Alignment, Font, PatternFill
from app.company_data.processor import *
import os

def generate_lep_report(id_company):

    name_file = 'Отчет_ЛЭП.xlsx'
    new_file = 'static/Отчет_ЛЭП_new.xlsx'
    # try:
    #     os.remove(new_file)
    # except:
    #     pass

    wb = openpyxl.load_workbook(name_file)
    sheet = wb.active
    cursor = 6
    # id_company = 1


    def add_netowrk_label(network, cursor):
        sheet.merge_cells(f'A{cursor}:M{cursor}')
        fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')
        sheet[f'A{cursor}'].alignment = Alignment(horizontal='center')
        sheet[f'A{cursor}'] = network.get("Название")
        sheet[f'A{cursor}'].fill = fill


    def to_float(value):
        try:
            return float(value)
        except:
            return float(0)


    def add_lep_label(network, cursor):
        leps = get_lep(network.get('id'))
        for lep in leps:
            provoda = get_provod(lep.get('id'))
            new_cursor = cursor
            qty_cep = 0
            length_tracca = 0
            length_one_cep = 0
            len_ych_tracca = 0
            len_ych_one_cep = 0
            for provod in provoda:
                sheet[f'K{new_cursor}'] = provod.get("Марка")
                sheet[f'E{new_cursor}'] = provod.get("Год_ввода") or '—'
                sheet[f'M{new_cursor}'] = 2019 - int(provod.get("Год_ввода", 0) or 0) if provod.get("Год_ввода") else '—'
                qty_cep += int(provod.get("Количество_цепей", 0) or 0)
                length_one_cep += to_float(provod.get("Длина_по_цепям", 0) or 0)
                len_ych_tracca += to_float(provod.get("Длина_участка_по_трассе", 0) or 0)
                len_ych_one_cep += to_float(provod.get("Длина_участка_по_цепям", 0) or 0)
                new_cursor += 1
            if new_cursor - 1 < cursor:
                new_cursor += 1
            sheet.merge_cells(f'B{cursor}:B{new_cursor - 1}')
            sheet[f'B{cursor}'] = lep.get("Диспетчерское_наименование")
            sheet.merge_cells(f'C{cursor}:C{new_cursor - 1}')
            sheet[f'C{cursor}'] = lep.get("Название")
            sheet.merge_cells(f'D{cursor}:D{new_cursor - 1}')
            sheet[f'D{cursor}'] = lep.get("Напряжение")
            sheet.merge_cells(f'F{cursor}:F{new_cursor - 1}')
            sheet[f'F{cursor}'] = qty_cep or 0
            sheet.merge_cells(f'G{cursor}:G{new_cursor - 1}')
            sheet[f'G{cursor}'] = length_tracca or '—'
            sheet.merge_cells(f'H{cursor}:H{new_cursor - 1}')
            sheet[f'H{cursor}'] = length_one_cep or '—'
            sheet.merge_cells(f'I{cursor}:I{new_cursor - 1}')
            sheet[f'I{cursor}'] = len_ych_tracca or '—'
            sheet.merge_cells(f'J{cursor}:J{new_cursor - 1}')
            sheet[f'J{cursor}'] = len_ych_one_cep or '—'
            sheet.merge_cells(f'L{cursor}:L{new_cursor - 1}')
            sheet[f'L{cursor}'] = lep.get("Техническое_состояние")
            cursor = new_cursor + 1
        return cursor


    def add_company_label(id_company, cursor):
        sql_query = f'select * from "Компания" where id = {id_company}'
        company = Sql.exec(query=sql_query)[0]
        sheet.merge_cells(f'A{cursor}:M{cursor}')
        fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        sheet[f'A{cursor}'].alignment = Alignment(horizontal='center')
        sheet[f'A{cursor}'] = company.get("Название")
        sheet[f'A{cursor}'].fill = fill



    networks = get_networks(id_company)

    add_company_label(1, cursor)
    sheets = [sheet]
    for network in range(len(networks) - 1):
        sheets.append(wb.copy_worksheet(sheet))
    cursor += 1
    for i, network in enumerate(networks):
        sheet = sheets[i]
        sheet.title = network.get('Название')
        add_netowrk_label(network, cursor)
        cursor += 1
        cursor = add_lep_label(network, cursor)
        cursor = 7


    wb.save(new_file)

# generate_lep_report(1)